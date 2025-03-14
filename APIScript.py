# ================================================================================================================================
# King County International Airport
# Department of Executive Services, Innovation Team
# Author: Collin Campagne
# Purpose: Uses an API (airportdb.io) to get XY coordinates of airports in Washington State. Uses ICAO codes 
#   within an input table to search API for corresponding codes, outputting an XY table with Airport Name, ICAO Code, Latitude and Longitude.
# Parameters: Code Table
# Dependencies: None
# Output: Spreadsheet in project folder containing airport name, ICAO code, latitude, and longitude of airports within source table. 
# Date Created: February 24, 2025
# Date Updated: 
# Editor: 
# ================================================================================================================================

# Import necessary libraries.
import requests
import os
from openpyxl import Workbook
from openpyxl import load_workbook
import arcpy
import traceback

# ================================================================================================================================

aprx = arcpy.mp.ArcGISProject("CURRENT")
current_map = aprx.activeMap
project_folder = aprx.homeFolder
output_path = os.path.join(project_folder, 'WAMA_LatLong.xlsx')
Spreadsheet = arcpy.GetParameterAsText(0)

# ================================================================================================================================

# Create Workbook if it doesn't exist
if not os.path.exists(output_path):
    workbook = Workbook()
    worksheet = workbook.active
    # Define the headers
    headers = ['Name', 'Code', 'Latitude', 'Longitude']
    # Write the headers to the first row
    worksheet.append(headers)
    workbook.save(output_path)
else:
    workbook = load_workbook(output_path)
    worksheet = workbook.active

# ================================================================================================================================

# create a function that gets data from API
# Source: https://airportdb.io/, which is sourced from https://ourairports.com/
def airportInfo(ICAO):
    endpoint = f'https://airportdb.io/api/v1/airport/{ICAO}?apiToken=[REDACTED]'
    try: 
        response = requests.get(endpoint)


        if response.status_code == 200:
            data = response.json() # make sure data is fed in JSON format
            airportName = data.get('name') 
            code = data.get('ident') # 'ident' is the name AirportDB uses for the ICAO Code 
            latitude = data.get('latitude_deg')
            longitude = data.get('longitude_deg')

            # Add messages so user can see if the data is being accessed properly
            arcpy.AddMessage(f"Airport Name: {airportName}")
            arcpy.AddMessage(f"ICAO Code: {code}")
            arcpy.AddMessage(f"Latitude: {latitude}")
            arcpy.AddMessage(f"Longitude: {longitude}")

            worksheet.append([airportName, code, latitude, longitude])
            workbook.save(output_path)
            arcpy.AddMessage(f"Info successfully written to {output_path}")

        # Error handling
        else: 
            arcpy.AddMessage(f"Failed to retrieve data for ICAO code: {ICAO}. Status Code: {response.status_code}")
    # More Error Handling
    except Exception as e:
        arcpy.AddMessage(f"Error fetching data for ICAO {ICAO}: {str(e)}")

# ================================================================================================================================

# Use another try/except method to use ICAO codes from input table to query API and write data to output spreadsheet
try:
    with arcpy.da.SearchCursor(Spreadsheet, ['CODE']) as cursor:
        for row in cursor:
            icaoCode = row[0]
            if isinstance(icaoCode, str):
                airportInfo(icaoCode.strip()) # 'Strip' removes any spaces or blanks that may erroneously be in the cell.
            else:
                arcpy.AddMessage(f"Invalid ICAO code: {icaoCode}")

# ================================================================================================================================

# Error Handling:
except arcpy.ExecuteError:
    arcpy.AddMessage("ArcPy Error: " + arcpy.GetMessages(2))

except Exception as e:
    tb = traceback.format_exc()
    arcpy.AddMessage(f"Python Error: {tb}")
